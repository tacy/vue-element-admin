import base64
import logging
import random
from datetime import date, timedelta

import arrow
from django.db import connection, transaction
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework import status, views
from rest_framework.response import Response

from stock.models import Order

YMTKEY = {
    '东京彩虹桥': {
        'appid': 'llzlHWWDTkEsUUjwKf',
        'appsecret': 'xdP5yraJQdpypKZNQ0M0zqE35dcrEWox',
        'authcode': 'Ul1BpFlBHdLR6EnEv75RV6QeradgjdBk'
    },
    '妈妈宝宝日本馆': {
        'appid': 'B9EBxjEN4JYB58BG4B',
        'appsecret': 'AKiwySBsiIwqz2TkkgQPXOJgCooc97Jt',
        'authcode': '6SJRmS03o6kwoYjqNPjUXocfMK0MpLhT'
    }
}

access_token = 'AESaZpmFNNcLRbNFmWK38S2ELvpzwjHkRjkpJkNmaaRIpEJ7T+FYBfVvoekui/2k1g=='
client_secret = 'APvYM8Mt5Xg1QYvker67VplTPQRx28Qt/XPdY9D7TUhaO3vgFWQ71CRZ/sLZYrn97w=='.lower(
)
client_id = '8417db83-360c-4275-974f-cf9a2734d8f8'
uex_user = '2830020@qq.com'
uex_passwd = '20162017'

# debug
# access_token = 'ACiYUZ6aKC48faYFD6MpvbOf73BdE9OV5g15q1A6Ghs+i/XIawq/9RHJCzc6Y3UNxA=='
# client_secret = 'APvYM8Mt5Xg1QYvker67VplTPQRx28Qt/XPdY9D7TUhaO3vgFWQ71CRZ/sLZYrn97w=='.lower(
# )
# client_id = '68993573-E38D-4A8A-A263-055C401F9369'

logger = logging.getLogger(__name__)


class ExportBondedOrder(views.APIView):
    def post(self, request, format=None):
        excel_data = [
            [
                '订单号', '订单时间', '收件人名称', '收件人电话', '收件人省', '收件人市', '收件人区',
                '收件人地址', '订购人证件号码', '商品海关备案货号', '申报数量', '物流企业'
            ],
        ]
        sql = "select o.id, o.orderid, o.piad_time, o.receiver_name, o.receiver_mobile, o.receiver_address, receiver_idcard, b.filing_no, o.quantity, '中通' from stock_order o inner join stock_bondedproduct b on o.jancode=b.jancode where o.status='待处理' and o.export_status is null and o.delivery_type='第三方保税' and o.channel_name='洋码头' and b.bonded_name='郑州保税'"
        results = []
        with connection.cursor() as c:
            c.execute(sql)
            results = c.fetchall()
        for r in results:
            addSplit = r[5].split(',')
            excel_data.append([
                r[1], r[2], r[3], r[4], addSplit[0], addSplit[1], addSplit[2],
                r[5].replace('#', '-'), r[6], r[7], r[8], r[9]
            ])

        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title='郑州保税订单')
        for line in excel_data:
            ws.append(line)

        # response = HttpResponse(
        #     content_type=
        #     'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        # )
        # response['Content-Disposition'] = 'attachment; filename=mydata.xlsx'
        # wb.save(response)
        base64Data = base64.b64encode(save_virtual_workbook(wb))
        msg = {'tableData': base64Data}

        with transaction.atomic():
            for r in results:
                Order.objects.filter(id=r[0]).update(
                    export_status='已导出',
                    status='已发货',
                    channel_delivery_status='已发货')
        return Response(data=msg, status=status.HTTP_200_OK)


class ExportDomesticOrder(views.APIView):
    def post(self, request, format=None):
        ords = request.data
        excel_data = [
            [
                '订单号', '收件人', '固话', '手机', '地址', '明细', '发件人', '发件人电话', '发件人地址',
                '备注', '代收金额', '保价金额', '业务类型'
            ],
        ]

        for o in ords:
            if '拼邮' not in o['shipping_name'] or '已发货' in o['status']:
                errmsg = {'errmsg': '订单:%s物流方式非拼邮或者状态为已发货' % (o['orderid'])}
                return Response(
                    data=errmsg, status=status.HTTP_400_BAD_REQUEST)
            excel_data.append([
                o['orderid'], o['receiver_name'], o['receiver_mobile'],
                o['receiver_mobile'], o['receiver_address'],
                o['product_title'], o['seller_name'], '13922442299', '', '',
                '', '', ''
            ])
        if excel_data:
            wb = Workbook(write_only=True)
            ws = wb.create_sheet(title='主表')
            for line in excel_data:
                ws.append(line)

        # response = HttpResponse(
        #     content_type=
        #     'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        # )
        # response['Content-Disposition'] = 'attachment; filename=mydata.xlsx'
        # wb.save(response)

        # https://stackoverflow.com/questions/8469665/saving-openpyxl-file-via-text-and-filestream
        base64Data = base64.b64encode(save_virtual_workbook(wb))
        msg = {'tableData': base64Data}
        with transaction.atomic():
            for o in ords:
                Order.objects.filter(id=o['id']).update(export_status='已导出')

        return Response(data=msg, status=status.HTTP_200_OK)


class ExportPrint(views.APIView):
    def post(self, request, format=None):
        ords = request.data
        excel_data = [
            ['物流单号', '订单号', '收件人', '地址', '产品', '电话', '发件人', '发件人地址', '发件人电话'],
        ]
        ids = [o['id'] for o in ords]
        for o in ords:
            orderObjs = Order.objects.filter(orderid=o['orderid'])
            oObj = orderObjs[0]
            if '轨迹' not in oObj.shipping.name or '待发货' not in oObj.status or not oObj.shippingdb:
                errmsg = {'errmsg': '订单:%s物流方式非轨迹或非待发货状态' % (o['orderid'])}
                return Response(
                    data=errmsg, status=status.HTTP_400_BAD_REQUEST)
            excel_data.append([
                oObj.shippingdb.db_number,
                oObj.orderid,
                oObj.receiver_name,
                oObj.receiver_address,
                oObj.product_title,
                oObj.receiver_mobile,
                '天狗',
                '東京都練馬区石神井台埼玉県朝霞市泉水3-7-9-115',
                '400-650-8988',
            ])
        if excel_data:
            wb = Workbook(write_only=True)
            ws = wb.create_sheet(title='热敏数据')
            for line in excel_data:
                ws.append(line)

        base64Data = base64.b64encode(save_virtual_workbook(wb))
        msg = {'tableData': base64Data}

        Order.objects.filter(id__in=ids).update(importstatus='已打印')
        return Response(data=msg, status=status.HTTP_200_OK)


class ExportUexTrack(views.APIView):
    def get(self, request, format=None):
        exportTemplete = {
            '日本海关': ['待导出', 1],
            '中国海关': ['日本海关', 4],
        }
        exportType = request.query_params.get('exportType')
        data = exportTemplete[exportType]
        excel_data = [
            ['物流信息', '时间', '6号库出库码'],
        ]
        ords = Order.objects.filter(
            shipping__name='轨迹',
            export_status=data[0],
            piad_time__lte=date.today() - timedelta(days=data[1]))

        # 生成轨迹
        ods = set([(o.piad_time, o.shippingdb.db_number)
                   for o in ords])  # 去重db_number
        length = len(ods)
        trace_time = []
        for idx, o in enumerate(ods):
            # 生成时间轨迹
            ordtime = arrow.get(o[0])
            if exportType == '日本海关':
                t1 = ordtime.replace(hour=0).shift(
                    days=+1, seconds=random.randint(36000, 39600))
                t2 = t1.replace(hour=0).shift(
                    days=+1, seconds=random.randint(32400, 36000))
                t3 = t2.replace(hour=0).shift(
                    seconds=random.randint(43200, 46800))
                t4 = t3.replace(hour=0).shift(
                    days=+1, seconds=random.randint(54000, 57600))
                trace_time.append((idx, '日本仓已接单',
                                   t1.format('YYYY-MM-DD HH:mm:ss'), o[1]))
                trace_time.append((idx + length, '生成单证信息',
                                   t2.format('YYYY-MM-DD HH:mm:ss'), o[1]))
                trace_time.append((idx + 2 * length, '离开日本仓',
                                   t3.format('YYYY-MM-DD HH:mm:ss'), o[1]))
                trace_time.append((idx + 3 * length, '日本海关申报中',
                                   t4.format('YYYY-MM-DD HH:mm:ss'), o[1]))
            else:
                t1 = ordtime.replace(hour=0).shift(
                    days=+3, seconds=random.randint(61200, 64800))
                t2 = t1.replace(hour=0).shift(
                    days=+1, seconds=random.randint(32400, 36000))
                t3 = t2.shift(seconds=random.randint(12600, 14400))
                t4 = t3.shift(seconds=random.randint(14400, 18000))
                t5 = t4.replace(hour=0).shift(
                    days=+2, seconds=random.randint(36000, 39600))
                trace_time.append((idx, '已获得出境许可',
                                   t1.format('YYYY-MM-DD HH:mm:ss'), o[1]))
                trace_time.append((idx + length, '航班已起飞，离开日本东京',
                                   t2.format('YYYY-MM-DD HH:mm:ss'), o[1]))
                trace_time.append((idx + 2 * length, '航班降落，到达广州机场',
                                   t3.format('YYYY-MM-DD HH:mm:ss'), o[1]))
                trace_time.append((idx + 3 * length, '到达机场快件中心',
                                   t4.format('YYYY-MM-DD HH:mm:ss'), o[1]))
                trace_time.append((idx + 4 * length, '海关清关中',
                                   t5.format('YYYY-MM-DD HH:mm:ss'), o[1]))

        # 添加轨迹
        for t in sorted(trace_time):
            excel_data.append(list(t[1:]))
        if excel_data:
            wb = Workbook(write_only=True)
            ws = wb.create_sheet(title=exportType)
            for line in excel_data:
                ws.append(line)

        base64Data = base64.b64encode(save_virtual_workbook(wb))
        msg = {'tableData': base64Data}
        ords.update(export_status=exportType)

        return Response(data=msg, status=status.HTTP_200_OK)


class CategoryGet(views.APIView):
    def get(self, request, format=None):
        sql = 'select b.category_id category_id, b.category_cn_name category_cn_name, a.category_id category_parent_id, a.category_cn_name category_parent_cn_name, a.category_version category_version from stock_category a join stock_category b on (a.category_id=b.category_parent_id)'

        def dictfetchall(cursor):
            "Return all rows from a cursor as a dict"
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

        with connection.cursor() as c:
            c.execute(sql)
            results = dictfetchall(c)
            relationData = {}
            for i in results:
                dictkey = i['category_parent_id'].zfill(5)
                if dictkey in relationData:
                    relationData[dictkey]['children'].append({
                        'label':
                        i['category_cn_name'],
                        'value':
                        i['category_id']
                    })
                else:
                    relationData[dictkey] = {
                        'label':
                        i['category_parent_cn_name'],
                        'value':
                        i['category_parent_id'],
                        'children': [
                            {
                                'label': i['category_cn_name'],
                                'value': i['category_id']
                            },
                        ]
                    }

            sortData = [relationData[i] for i in sorted(relationData.keys())]
            data = {
                'results': sortData,
            }
            return Response(data=data, status=status.HTTP_200_OK)
        return Response(status=status.HTTP_400_BAD_REQUEST)
