import base64
import asyncio
import json
import logging
from collections import OrderedDict
from django.http import HttpResponse
from io import BytesIO

import arrow.arrow
import aiohttp
import django_filters.rest_framework
from PyPDF2 import PdfFileMerger
from django.db import IntegrityError, connection, transaction
from django.db.models import F, Max
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework import generics, status, views
from rest_framework.authtoken.models import Token
from rest_framework.response import Response

from .filter import (OrderFilter, ProductFilter, PurchaseOrderFilter,
                     ShippingDBFilter, StockFilter, PurchaseOrderItemFilter)
from .models import (BondedProduct, Inventory, Order, Product, PurchaseOrder,
                     PurchaseOrderItem, Shipping, ShippingDB, Stock, Supplier)
from .serializers import (BondedProductSerializer, InventorySerializer,
                          OrderSerializer, ProductSerializer,
                          PurchaseOrderItemSerializer, PurchaseOrderSerializer,
                          ShippingDBSerializer, ShippingSerializer,
                          StockSerializer, SupplierSerializer, TokenSerializer)
from ymatou import uex, utils, ymatouapi

YMTKEY = {
    '东京彩虹桥': {
        'appid': 'llzlHWWDTkEsUUjwKf',
        'appsecret': 'xdP5yraJQdpypKZNQ0M0zqE35dcrEWox',
        'authcode': 'Ul1BpFlBHdLR6EnEv75RV6QeradgjdBk'
    },
    '妈妈宝宝日本馆': {
        'appid': 'B9EBxjEN4JYB58BG4B',
        'appsecret': 'AKiwySBsiIwqz2TkkgQPXOJgCooc97Jt',
        'authcode': '6SJRmS03o6kwoYjqNPjUXocfMK0MpLhT'
    }
}

access_token = 'AESaZpmFNNcLRbNFmWK38S2ELvpzwjHkRjkpJkNmaaRIpEJ7T+FYBfVvoekui/2k1g=='
client_secret = 'APvYM8Mt5Xg1QYvker67VplTPQRx28Qt/XPdY9D7TUhaO3vgFWQ71CRZ/sLZYrn97w=='.lower(
)
client_id = '8417db83-360c-4275-974f-cf9a2734d8f8'
uex_user = '2830020@qq.com'
uex_passwd = '20162017'

# debug
# access_token = 'ACiYUZ6aKC48faYFD6MpvbOf73BdE9OV5g15q1A6Ghs+i/XIawq/9RHJCzc6Y3UNxA=='
# client_secret = 'APvYM8Mt5Xg1QYvker67VplTPQRx28Qt/XPdY9D7TUhaO3vgFWQ71CRZ/sLZYrn97w=='.lower(
# )
# client_id = '68993573-E38D-4A8A-A263-055C401F9369'

logger = logging.getLogger(__name__)


# 1. 生成DB面单, 如果是码头订单, 需要先确认订单状态, 状态异常, 直接返回异常
# 2. 需要考虑该用户是否有其他订单, 如果有其他订单, 需要提醒操作人员.
class XloboCreateNoVerification(views.APIView):
    def post(self, request, format=None):
        data = request.data
        ords = data['orders']

        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop = asyncio.get_event_loop()
        sess = aiohttp.ClientSession(loop=loop)

        # ymatou order need check orderstatus
        # 这里可能有合并订单发货情况, 需要根据订单ID去重, 然后去码头后台查每一
        # 个订单状态是否正常
        if '洋码头' in ords[0]['channel_name']:
            skey = YMTKEY[ords[0]['seller_name']]
            ymtapi = ymatouapi.YmatouAPI(sess, skey['appid'],
                                         skey['appsecret'], skey['authcode'])
            ordids = list([OrderedDict.fromkeys([i['orderid'] for i in ords])])
            for oid in ordids:
                result = loop.run_until_complete(
                    ymtapi.getOrderInfo(ords[0]['orderid']))
                # result = loop.run_until_complete(ymtapi.getOrderInfo('127086025'))
                for oi in result['content']['order_info']['order_items_info']:
                    if oi['refund_id'] == 0 or oi['refund_id'] == 1:
                        errmsg = {'errmsg': '订单状态异常, 请到码头后台确认'}
                        return Response(
                            data=errmsg, status=status.HTTP_400_BAD_REQUEST)

        # 如果need_check, 需要先查一下该用户是否有其他订单没有一并提交, 如果有, 返回
        # 异常
        if not data['disable_check']:
            check_ords = Order.objects.filter(
                receiver_name=ords[0]['receiver_name'],
                receiver_mobile=ords[0]['receiver_mobile'],
                shippingdb__isnull=True,
                status__in=['待处理', '待采购', '待发货', '已采购', '需介入'])
            if len(check_ords) != len(ords):
                errmsg = {'errmsg': '该用户有其他订单, 请检查.'}
                return Response(
                    data=errmsg, status=status.HTTP_400_BAD_REQUEST)

        # create db number
        # construct api msg
        channel_name = ords[0]['channel_name']
        address = ords[0]['receiver_address'].split(',')
        order_piad_time = ords[0]['piad_time']
        billSenderInfo = {
            'Name': ords[0]['seller_name'],
            'Address': '中央区新富1-3-15　京橋プリズムビル　４階',
            'Phone': '11',
        }
        billReceiverInfo = {
            'Name': ords[0]['receiver_name'],
            'Province': address[0],
            'City': address[1],
            'District': address[2],
            'Address': address[3],
            'Phone': ords[0]['receiver_mobile'],
            'PostCode': ords[0]['receiver_zip'],
            'IdCode': ords[0].get('receiver_idcard')
        }
        billSupplyInfo = {
            'OrderCode': ','.join(set([o['orderid'] for o in ords])),
            'TradingNo': '11111111',
            'ChannelName': channel_name
        }
        billCategoryList = []
        for o in ords:
            productObj = Product.objects.get(jancode=o['jancode'])
            bcl = {
                'CategoryId': productObj.category.id,
                'CategoryVersion': productObj.category.category_version,
                'Count': o['quantity'],
                'UnitPrice': o['price'],
                'ProductName': productObj.name,
                'Brand': productObj.brand,
                'Specification': productObj.specification
            }
            billCategoryList.append(bcl)

        data.pop('orders')
        data['BillSenderInfo'] = billSenderInfo
        data['BillReceiverInfo'] = billReceiverInfo
        data['BillSupplyInfo'] = billSupplyInfo
        data['BillCategoryList'] = billCategoryList

        xloboapi = ymatouapi.XloboAPI(sess, access_token, client_secret,
                                      client_id)
        result = loop.run_until_complete(xloboapi.createNoVerification(data))
        logger.debug('XloboCreateNoVerification: %s', result)
        if result['ErrorCount'] > 0:
            errmsg = {'errmsg': result['ErrorInfoList'][0]['ErrorDescription']}
            return Response(data=errmsg, status=status.HTTP_400_BAD_REQUEST)
        loop.close()
        # if result['ErrorCount']:
        #     return Response(data=result, status=status.HTTP_400_BAD_REQUEST)

        # save shippingdb & update order(shippingdb)
        with transaction.atomic():
            shippingObj = Shipping.objects.get(id=ords[0]['shipping'])
            inventoryObj = Inventory.objects.get(id=ords[0]['inventory'])
            shippingdbObj = ShippingDB(
                db_number=result['Result']['BillCode'],
                status='待处理',
                order_piad_time=order_piad_time,
                channel_name=channel_name,
                shipping=shippingObj,
                inventory=inventoryObj)
            shippingdbObj.save()
            for o in ords:
                orderObj = Order.objects.get(id=o['id'])
                orderObj.shippingdb = shippingdbObj
                orderObj.save(update_fields=['shippingdb'])

        return Response(data=result, status=status.HTTP_200_OK)


# 虚仓电商无需我们自己再处理, 直接贝海负责打包发货
class XloboCreateFBXBill(views.APIView):
    def post(self, request, format=None):
        data = request.data
        ords = data['orders']
        channinfo = {
            '洋码头': 2,
            '天猫': 3,
            '苏宁': 4,
            '京东': 5,
            '淘宝': 6,
            '一号店': 8,
            '当当': 9,
            'Higo': 10,
            '其他渠道': 7
        }
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop = asyncio.get_event_loop()
        sess = aiohttp.ClientSession(loop=loop)

        # ymatou order need check orderstatus
        # 这里可能有合并订单发货情况, 需要根据订单ID去重, 然后去码头后台查每一
        # 个订单状态是否正常
        if '洋码头' in ords[0]['channel_name']:
            skey = YMTKEY[ords[0]['seller_name']]
            ymtapi = ymatouapi.YmatouAPI(sess, skey['appid'],
                                         skey['appsecret'], skey['authcode'])
            ordids = list([OrderedDict.fromkeys([i['orderid'] for i in ords])])
            for oid in ordids:
                result = loop.run_until_complete(
                    ymtapi.getOrderInfo(ords[0]['orderid']))
                # result = loop.run_until_complete(ymtapi.getOrderInfo('127086025'))
                for oi in result['content']['order_info']['order_items_info']:
                    if oi['refund_id'] == 0 or oi['refund_id'] == 1:
                        errmsg = {'errmsg': '订单状态异常, 请到码头后台确认'}
                        return Response(
                            data=errmsg, status=status.HTTP_400_BAD_REQUEST)

        # 如果need_check, 需要先查一下该用户是否有其他订单没有一并提交, 如果有, 返回
        # 异常
        if not data['disable_check']:
            check_ords = Order.objects.filter(
                receiver_name=ords[0]['receiver_name'],
                receiver_mobile=ords[0]['receiver_mobile'],
                shippingdb__isnull=True,
                status__in=['待处理', '待采购', '待发货', '已采购', '需介入'])
            if len(check_ords) != len(ords):
                errmsg = {'errmsg': '该用户有其他订单, 请检查.'}
                return Response(
                    data=errmsg, status=status.HTTP_400_BAD_REQUEST)

        # construct api msg
        channel_name = ords[0]['channel_name']
        address = ords[0]['receiver_address'].split(',')
        order_piad_time = ords[0]['piad_time']
        billSenderInfo = {
            'Name': ords[0]['seller_name'],
            'Address': '中央区新富1-3-15　京橋プリズムビル　４階',
            'Phone': '11',
        }
        billReceiverInfo = {
            'Name': ords[0]['receiver_name'],
            'Province': address[0],
            'City': address[1],
            'District': address[2],
            'Address': address[3],
            'Phone': ords[0]['receiver_mobile'],
            'PostCode': ords[0]['receiver_zip'],
            'IdCode': ords[0].get('receiver_idcard')
        }
        billSupplyInfo = {
            'OrderCode': ords[0]['orderid'],
            'BillThirdPartType': channinfo[channel_name],
            'ChannelName': channel_name
        }
        billCategoryList = []
        for o in ords:
            bcl = {
                'ProductNo': o['jancode'],
                'Count': o['quantity'],
                'Price': o['price'],
            }
            billCategoryList.append(bcl)

        data.pop('orders')
        data['IsReceiveTax'] = data['IsRecTax']
        data['PackingType'] = data['IsRePacking']
        data['BillSenderInfo'] = billSenderInfo
        data['BillReceiverInfo'] = billReceiverInfo
        data['BillSupplyInfo'] = billSupplyInfo
        data['GoodsSkuInfos'] = billCategoryList

        xloboapi = ymatouapi.XloboAPI(sess, access_token, client_secret,
                                      client_id)
        result = loop.run_until_complete(xloboapi.createFBXBill(data))
        logger.debug('XloboCreateFBXBill: %s', result)
        if result['ErrorCount'] > 0:
            errmsg = {'errmsg': result['ErrorInfoList'][0]['ErrorDescription']}
            return Response(data=errmsg, status=status.HTTP_400_BAD_REQUEST)
        loop.close()
        # if result['ErrorCount']:
        #     return Response(data=result, status=status.HTTP_400_BAD_REQUEST)

        # save shippingdb & update order(shippingdb)
        with transaction.atomic():
            shippingObj = Shipping.objects.get(id=ords[0]['shipping'])
            inventoryObj = Inventory.objects.get(id=ords[0]['inventory'])
            shippingdbObj = ShippingDB(
                db_number=result['Result']['BillCode'],
                status='已出库',
                order_piad_time=order_piad_time,
                channel_name=channel_name,
                shipping=shippingObj,
                inventory=inventoryObj)
            shippingdbObj.save()
            for o in ords:
                orderObj = Order.objects.get(id=o['id'])
                orderObj.shippingdb = shippingdbObj
                orderObj.status = '已发货'  # 订单直接进入已发货状态
                orderObj.save(update_fields=['status', 'shippingdb'])
                stockObj = Stock.objects.get(
                    product__jancode=orderObj.jancode,
                    inventory=orderObj.inventory)
                stockObj.quantity = F('quantity') - orderObj.quantity  # 扣减库存
                stockObj.preallocation = F('preallocation') - orderObj.quantity
                stockObj.save()

        return Response(data=result, status=status.HTTP_200_OK)


# 删除DB面单, 清理订单的shipping_id字段
class XloboDeleteDBNumber(views.APIView):
    def post(self, request, format=None):
        # construct api msg
        data = request.data
        db_number = data['db_number']
        if data['shipping_name'] in ['虚仓电商', '直邮电商', '贝海直邮电商']:
            msg = {'BillCode': db_number}
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            sess = aiohttp.ClientSession(loop=loop)
            xloboapi = ymatouapi.XloboAPI(sess, access_token, client_secret,
                                          client_id)
            result = loop.run_until_complete(xloboapi.deleteDBNumber(msg))
            loop.close()
            logger.debug('XloboDeleteDBNumber: %s', result)
            if result['Result']['ErrorCount'] > 0:
                errmsg = {
                    'errmsg':
                    result['Result']['ErrorInfoList'][0]['ErrorDescription']
                }
                return Response(
                    data=errmsg, status=status.HTTP_400_BAD_REQUEST)
        with transaction.atomic():
            shippingdbObj = ShippingDB.objects.get(id=data['id'])
            shippingdbObj.status = '已删除'
            shippingdbObj.order.all().update(shippingdb=None)
            shippingdbObj.save(update_fields=[
                'status',
            ])

        return Response(status=status.HTTP_200_OK)


class XloboGetPDF(views.APIView):
    def get(self, request, format=None):
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop = asyncio.get_event_loop()
        # construct api msg
        # drf can't get query list
        # data = {'BillCodes': request.query_params.get('BillCodes[]')}
        data = {'BillCodes': self.request.GET.getlist("BillCodes[]")}
        # data = {
        #     'BillCodes': [
        #         'DB273208811JP',
        #     ]
        # }

        sess = aiohttp.ClientSession(loop=loop)
        xloboapi = ymatouapi.XloboAPI(sess, access_token, client_secret,
                                      client_id)
        result = loop.run_until_complete(xloboapi.getPDF(data))
        loop.close()
        logger.debug('XloboGetPDF: %s', result)
        if result['ErrorCount'] > 0:
            info = result['ErrorInfoList'][0]
            errmsg = {
                'errmsg': info['Identity'] + ':' + info['ErrorDescription']
            }
            return Response(data=errmsg, status=status.HTTP_400_BAD_REQUEST)
        merger = PdfFileMerger()
        pdftool = utils.PDFTool()
        sql = 'select p.name, p.specification, o.jancode, o.quantity, s.location from stock_order o inner join stock_product p on o.jancode=p.jancode inner join stock_stock s on s.product_id=p.id where o.shippingdb_id=%s and s.inventory_id=%s'
        for i, b in enumerate(result['Result']):
            db_bytes = base64.b64decode(b['BillPdfLabel'])
            ordsData = None

            with connection.cursor() as c:
                shippingdbObj = ShippingDB.objects.get(db_number=b['BillCode'])
                c.execute(sql, (shippingdbObj.id, shippingdbObj.inventory.id))
                # c.execute(sql, (8, ))
                ordsData = c.cursor.fetchall()
            # inp2 = BytesIO(pdftool.createShippingPDF(b['BillCode'], ordsData))
            # inp2 = open(
            #     '/home/tacy/workspace/python/lelewu/ymatou/simple_table.pdf',
            #     'rb')
            if not i:
                merger.append(fileobj=BytesIO(db_bytes), pages=(0, 1))
            else:
                merger.merge(
                    position=i * 2 + 1,
                    fileobj=BytesIO(db_bytes),
                    pages=(0, 1))
            merger.merge(
                position=2**(i + 1),
                fileobj=BytesIO(
                    pdftool.createShippingPDF(b['BillCode'], ordsData)),
                pages=(0, 1))

        res = BytesIO()
        merger.write(res)
        result['Result'] = [
            {
                'BillPdfLabel': base64.b64encode(res.getvalue())
            },
        ]
        res.close()

        # update shippingdb print_status
        ShippingDB.objects.filter(db_number__in=data['BillCodes']).update(
            print_status='已打印')

        return Response(data=result, status=status.HTTP_200_OK)
        # response = HttpResponse(content_type='application/pdf')
        # # response[
        # #     'Content-Disposition'] = 'attachment; filename="somefilename.pdf"'
        # response['Content-Disposition'] = 'inline; filename="somefilename.pdf"'
        # pdf = res.getvalue()
        # res.close()
        # response.write(pdf)
        # return response


class LogisticGet(views.APIView):
    def get(self, request, format=None):
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop = asyncio.get_event_loop()

        sess = aiohttp.ClientSession(loop=loop)
        xloboapi = ymatouapi.XloboAPI(sess, access_token, client_secret,
                                      client_id)
        result = loop.run_until_complete(xloboapi.getLogistic())
        loop.close()
        data = [{
            'value': i['LogisticId'],
            'key': i['LogisticName']
        } for i in result['Result']['LogisticInfoList']
                if i['iCountryId'] == 6]
        return Response(data=data, status=status.HTTP_200_OK)


class ExportBondedOrder(views.APIView):
    def get(self, request, format=None):
        excel_data = [
            ['header1', 'header2', 'header3', 'header4', 'header5'],
            [1, 4, 5, 6, 7],
            [5, 6, 2, 4, 8],
        ]

        if excel_data:
            wb = Workbook(write_only=True)
            ws = wb.create_sheet(title='主表')
            ws2 = wb.create_sheet(title='从表')
            for line in excel_data:
                ws.append(line)
                ws2.append(line)

        response = HttpResponse(
            content_type=
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=mydata.xlsx'

        wb.save(response)

        return response


class ExportDomesticOrder(views.APIView):
    def post(self, request, format=None):
        ords = request.data
        excel_data = [
            [
                '订单号',
                '收件人',
                '固话',
                '手机',
                '地址',
                '明细',
                '发件人',
                '发件人电话',
                '发件人地址',
                '备注',
                '代收金额',
                '保价金额',
                '业务类型',
            ],
        ]

        for o in ords:
            if '拼邮' not in o['shipping_name'] or '已发货' in o['status']:
                errmsg = {'errmsg': '订单:%s物流方式非拼邮或者状态为已发货' % (o['orderid'])}
                return Response(
                    data=errmsg, status=status.HTTP_400_BAD_REQUEST)
            excel_data.append([
                o['orderid'],
                o['receiver_name'],
                o['receiver_mobile'],
                o['receiver_mobile'],
                o['receiver_address'],
                o['product_title'],
                o['seller_name'],
                '13922442299',
                '',
                '',
                '',
                '',
                '',
            ])
        if excel_data:
            wb = Workbook(write_only=True)
            ws = wb.create_sheet(title='主表')
            for line in excel_data:
                ws.append(line)

        # response = HttpResponse(
        #     content_type=
        #     'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        # )
        # response['Content-Disposition'] = 'attachment; filename=mydata.xlsx'
        # wb.save(response)

        # https://stackoverflow.com/questions/8469665/saving-openpyxl-file-via-text-and-filestream
        base64Data = base64.b64encode(save_virtual_workbook(wb))
        msg = {'tableData': base64Data}
        with transaction.atomic():
            for o in ords:
                Order.objects.filter(id=o['id']).update(export_status='已导出')

        return Response(data=msg, status=status.HTTP_200_OK)


class SyncStock(views.APIView):
    def post(self, request, format=None):
        data = request.data
        impInventory = data['inventory_name']
        inv_dict = {
            '贝海': ['virtualstock-products', 'stock'],
            '广州': [u'国内库存出入库流水.xls', u'库存表'],
            '东京': ['tokyo_stock.xlsx', 'new'],
        }
        gsp = inv_dict[impInventory][0]
        wks = inv_dict[impInventory][1]
        inventoryObj = Inventory.objects.get(name=impInventory)
        syncer = utils.GoogleSpread()
        syncer.open_google_doc(gsp)
        stocks = syncer.read_google_doc_by_range(wks, 'A2:C', all_rows=True)
        with transaction.atomic():
            for i in list(syncer.chunks(stocks, 3)):
                if not i[2]:
                    continue
                productObj = None
                try:
                    productObj = Product.objects.get(jancode=i[0])
                except Product.DoesNotExist:
                    logger.info('Sync stock error, Jancode %s', i[0])
                    continue
                try:
                    stockObj = Stock.objects.get(
                        product=productObj, inventory=inventoryObj)
                    if '东京' in impInventory:
                        stockObj.quantity = F('quantity') + i[2]
                    else:
                        stockObj.quantity = i[2]
                    stockObj.save(update_fields=['quantity'])
                except Stock.DoesNotExist:
                    stockObj = Stock(
                        product=productObj,
                        quantity=i[2],
                        inventory=inventoryObj,
                        preallocation=0,
                        inflight=0, )
                    stockObj.save()
            return Response(status=status.HTTP_200_OK)
        return Response(status=status.HTTP_400_BAD_REQUEST)


class UexStockOut(views.APIView):
    def post(self, request, format=None):
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)

        data = request.data
        ords = data['orders']
        channel_name = ords[0]['channel_name']
        address = ords[0]['receiver_address'].split(',')
        # construct api msg
        payload = {
            'ship_id': data['ship_id'],
            'add_server[0]': 1,
            'user_order_no': ords[0]['orderid'],
            'send_user': ords[0]['seller_name'],
            'address[consignee]': ords[0]['receiver_name'],
            'address[phone]': ords[0]['receiver_mobile'],
            'address[card]': ords[0]['receiver_idcard'],
            'address[province]': address[0],
            'address[city]': address[1],
            'address[district]': address[2],
            'address[address]': address[3],
        }
        for i, v in enumerate(ords):
            payload['goods[' + str(i) + '][jan_code]'] = v['jancode']
            payload['goods[' + str(i) + '][num]'] = v['quantity']

        sess = aiohttp.ClientSession(loop=loop)
        uexapi = uex.UexAPI(sess, uex_user, uex_passwd)
        result = loop.run_until_complete(uexapi.login())
        result = loop.run_until_complete(uexapi.stockOut(payload))
        loop.close()
        logger.debug('UexStockOut: %s', result)
        result = json.loads(result)
        if not result['code']:
            errmsg = {'errmsg': result['msg']}
            return Response(data=errmsg, status=status.HTTP_400_BAD_REQUEST)
        with transaction.atomic():
            shippingObj = Shipping.objects.get(id=ords[0]['shipping'])
            inventoryObj = Inventory.objects.get(id=ords[0]['inventory'])
            shippingdbObj = ShippingDB(
                db_number=result['order_sn'],
                status='已出库',
                channel_name=channel_name,
                shipping=shippingObj,
                inventory=inventoryObj)
            shippingdbObj.save()
            for o in ords:
                orderObj = Order.objects.get(id=o['id'])
                orderObj.shippingdb = shippingdbObj
                orderObj.status = '已发货'  # 订单直接进入已发货状态
                orderObj.save(update_fields=['status', 'shippingdb'])
                stockObj = Stock.objects.get(
                    product__jancode=orderObj.jancode,
                    inventory=orderObj.inventory)
                stockObj.quantity = F('quantity') - orderObj.quantity  # 扣减库存
                stockObj.preallocation = F('preallocation') - orderObj.quantity
                stockObj.save()
        return Response(data=result, status=status.HTTP_200_OK)


# 思考: 拼邮订单, 还是需要填写正确的EMS单号, 否则不容易追踪包裹情况, 但是存在
# 不正确填写EMS单号的情况, 这个需要怎么处理?
class ManualAllocateDBNumber(views.APIView):
    def post(self, request, format=None):
        ords = request.data['orders']

        # ymatou order need check orderstatus
        if '洋码头' in ords[0]['channel_name']:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop = asyncio.get_event_loop()
            sess = aiohttp.ClientSession(loop=loop)

            skey = YMTKEY[ords[0]['seller_name']]
            ymtapi = ymatouapi.YmatouAPI(sess, skey['appid'],
                                         skey['appsecret'], skey['authcode'])
            result = loop.run_until_complete(
                ymtapi.getOrderInfo(ords[0]['orderid']))
            # result = loop.run_until_complete(ymtapi.getOrderInfo('127086025'))
            for oi in result['content']['order_info']['order_items_info']:
                if oi['refund_id'] == 0 or oi['refund_id'] == 1:
                    errmsg = {'errmsg': '订单状态异常, 请到码头后台确认'}
                    return Response(
                        data=errmsg, status=status.HTTP_400_BAD_REQUEST)

        db_number = request.data['Comment']
        channel_name = ords[0]['channel_name']
        order_piad_time = ords[0]['piad_time']
        # delivery_type = ords[0]['delivery_type']
        with transaction.atomic():
            orderStatus = None
            shippingdbObj = None
            try:
                shippingdbObj = ShippingDB.objects.get(db_number=db_number)
                for o in ords:
                    if '拼邮' not in o['shipping_name']:
                        errmsg = {'errmsg': '非拼邮订单, 面单号被重复使用, 请仔细检查确认'}
                        return Response(
                            data=errmsg, status=status.HTTP_400_BAD_REQUEST)
            except ShippingDB.DoesNotExist:
                shippingObj = Shipping.objects.get(id=ords[0]['shipping'])
                inventoryObj = Inventory.objects.get(id=ords[0]['inventory'])
                dbStatus = '待处理'
                if '贝海' in inventoryObj.name and 'EMS' in shippingObj.name:  # 贝海EMS无需我们打包
                    dbStatus = '已出库'
                    orderStatus = '已发货'  # TODO: 采购可能没有到库
                if '拼邮' in shippingObj.name:  # 拼邮不进入待发货列表, 但是需打包发货
                    dbStatus = '已出库'
                shippingdbObj = ShippingDB(
                    db_number=db_number,
                    # status='已出库' if '拼邮' in delivery_type else '待处理',   # 如果是拼邮订单, 只是需要一个面单回填, 无需做国外发货处理
                    status=dbStatus,
                    channel_name=channel_name,
                    order_piad_time=order_piad_time,
                    shipping=shippingObj,
                    inventory=inventoryObj)
                shippingdbObj.save()
            for o in ords:
                orderObj = Order.objects.get(id=o['id'])
                orderObj.shippingdb = shippingdbObj
                if orderStatus:
                    orderObj.status = orderStatus
                    stockObj = Stock.objects.get(
                        product__jancode=orderObj.jancode,
                        inventory=orderObj.inventory)
                    stockObj.quantity = F(
                        'quantity') - orderObj.quantity  # 扣减库存
                    stockObj.preallocation = F(
                        'preallocation') - orderObj.quantity
                    stockObj.save()
                orderObj.save(update_fields=['shippingdb', 'status'])

        return Response(status=status.HTTP_200_OK)


# 订单发货
class StockOut(views.APIView):
    # 1. 标记db面单状态, 设置运单号(delivery_no);
    # 2. 标记订单状态, 过滤条件是订单状态: ('待发货')
    # 3. 扣减库存
    def post(self, request, format=None):
        delivery_no = request.data['delivery_no']
        dbs = request.data['db_numbers'].split('\n')
        results = None
        try:
            with transaction.atomic():
                for db in dbs:
                    shippingdbObj = ShippingDB.objects.get(db_number=db)
                    if '已出库' in shippingdbObj.status:
                        results = {
                            'errmsg':
                            '面单:{} 已出库, 运单号:{}, 请确认订单是否已发货'.format(
                                db, shippingdbObj.delivery_no)
                        }
                        if shippingdbObj.delivery_no == delivery_no:
                            results = {
                                'errmsg': '面单:{} 重复录入或重复打包, 请检查'.format(db)
                            }
                        raise IntegrityError
                    shippingdbObj.status = '已出库'
                    shippingdbObj.delivery_no = delivery_no
                    shippingdbObj.delivery_time = arrow.now().format(
                        'YYYY-MM-DD HH:mm:ss')
                    shippingdbObj.save(update_fields=[
                        'status', 'delivery_no', 'delivery_time'
                    ])
                    for o in shippingdbObj.order.filter(
                            status__in=['待发货', '已采购']):
                        if '已采购' in o.status:
                            results = {
                                'errmsg':
                                '面单{}对应的订单:{}, 采购在途, 采购单号:{}, 请确认'.format(
                                    db, o.orderid, o.purchaseorder.orderid)
                            }
                            raise IntegrityError
                        o.status = '已发货'
                        o.save(update_fields=['status'])
                        stockObj = Stock.objects.get(
                            product__jancode=o.jancode, inventory=o.inventory)
                        stockObj.quantity = F('quantity') - o.quantity
                        stockObj.preallocation = F('preallocation') - o.quantity
                        stockObj.save()
        except IntegrityError:
            return Response(data=results, status=status.HTTP_400_BAD_REQUEST)

        return Response(status=status.HTTP_200_OK)


# 拼邮订单和第三方保税订单出库操作
# 更新订单状态, 扣减库存
class OrderOut(views.APIView):
    def post(self, request, format=None):
        ord = request.data
        with transaction.atomic():
            ordObjs = Order.objects.get(id=ord['id'])
            if '已发货' not in ordObjs.status:
                ordObjs.status = '已发货'
                ordObjs.save(update_fields=['status'])
                productObj = Product.objects.get(jancode=ordObjs.jancode)
                stockObj = Stock.objects.get(
                    product=productObj, inventory=ordObjs.inventory)
                stockObj.preallocation = F('preallocation') - ordObjs.quantity
                stockObj.quantity = F('quantity') - ordObjs.quantity
                stockObj.save()
                return Response(status=status.HTTP_200_OK)
            else:
                results = {'errmsg': '订单已发货'}
                return Response(
                    data=results, status=status.HTTP_400_BAD_REQUEST)


# 不要使用, 有问题, stock没有jancode字段了
class OrderItemGet(views.APIView):
    def get(self, request, format=None):
        sql = (
            'select o.orderid orderid, o.jancode jancode, o.product_title product_title, '
            'o.sku_properties_name sku_properties_name, o.quantity quantity, '
            'o.receiver_name receiver_name, o.receiver_address receiver_address, '
            'o.receiver_mobile receiver_mobile, s.location location, sdb.db_number db_number '
            'from stock_order o inner join stock_stock s '
            'on o.jancode=s.jancode and o.inventory_id=s.inventory_id and o.shippingdb_id=%s '
            'inner join stock_shippingdb sdb on o.shippingdb_id=sdb.id')

        def dictfetchall(cursor):
            "Return all rows from a cursor as a dict"
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

        db_number = request.query_params.get('shippingdb_id')
        with connection.cursor() as c:
            c.execute(sql, (db_number, ))
            results = dictfetchall(c)
            data = {
                'results': results,
            }
            return Response(data=data, status=status.HTTP_200_OK)
        return Response(status=status.HTTP_400_BAD_REQUEST)


class CategoryGet(views.APIView):
    def get(self, request, format=None):
        sql = 'select b.category_id category_id, b.category_cn_name category_cn_name, a.category_id category_parent_id, a.category_cn_name category_parent_cn_name, a.category_version category_version from stock_category a join stock_category b on (a.category_id=b.category_parent_id)'

        def dictfetchall(cursor):
            "Return all rows from a cursor as a dict"
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

        with connection.cursor() as c:
            c.execute(sql)
            results = dictfetchall(c)
            relationData = {}
            for i in results:
                dictkey = i['category_parent_id'].zfill(5)
                if dictkey in relationData:
                    relationData[dictkey]['children'].append({
                        'label':
                        i['category_cn_name'],
                        'value':
                        i['category_id']
                    })
                else:
                    relationData[dictkey] = {
                        'label':
                        i['category_parent_cn_name'],
                        'value':
                        i['category_parent_id'],
                        'children': [
                            {
                                'label': i['category_cn_name'],
                                'value': i['category_id']
                            },
                        ]
                    }

            sortData = [relationData[i] for i in sorted(relationData.keys())]
            data = {
                'results': sortData,
            }
            return Response(data=data, status=status.HTTP_200_OK)
        return Response(status=status.HTTP_400_BAD_REQUEST)


# 录入订单
class OrderTPRCreate(views.APIView):
    def put(self, request, format=None):
        data = request.data
        t = arrow.now()
        piad_time = t.format('YYYY-MM-DD HH:mm:ss')
        # orderid = 'T' + t.format('YYMMDD') + ''.join(
        #     random.choices(string.digits, k=2))
        with transaction.atomic():
            for p in data['products']:
                o = {
                    'orderid': data['orderid'],
                    'piad_time': piad_time,
                    'delivery_type': data['delivery_type'],
                    'seller_name': data['seller_name'],
                    'channel_name': data['channel_name'],
                    'receiver_name': data['receiver_name'],
                    'receiver_address': data['receiver_address'],
                    'receiver_zip': data.get('receiver_zip'),
                    'receiver_mobile': data['receiver_mobile'],
                    'receiver_idcard': data.get('receiver_idcard'),
                    'seller_memo': data.get('seller_memo'),
                    'jancode': p['jancode'],
                    'product_title': p['product_title'],
                    'sku_properties_name': p['sku_properties_name'],
                    'price': p['price'],
                    'payment': float(p['price']) * float(p['quantity']),
                    'quantity': p['quantity']
                }
                orderObj = Order(**o)
                orderObj.save()
        return Response(status=status.HTTP_200_OK)


# 订单预处理
class OrderAllocate(views.APIView):
    # 派单涉及到两个表: order和stock
    # 派单需要的操作: 1. 占用库存(preallocation); 2. 标记订单需要采购的数量(need_purchase);
    # 3. 修改订单状态(status:待发货/待采购)
    #
    # 派单流程: 派单分为派单和重派, 如果订单的inventory字段为空, 定位为派单;
    # 如果该字段非空, 定义为重派.
    #
    # 计算商品库存是否能满足订单发需求的时候, 需要特别注意, 如果在库实际不足, 但是在库+在途能满足,
    # 需要标记订单状态为已采购, 同时关联最晚采购该商品的采购订单
    #
    # 需要根据传入订单号, 查询数据库表中对应订单, 判断order_inventory字段内容:
    # if paramorder.status == '已删除'; then delete order opetion, use in conflict
    # if paramorder.inventory is null; then return
    # if dborder.inventory is null; then 派单
    # if dborder.inventory != paramorder.inventory; then 重派
    # if dborder.inventory == paramorder.inventory:
    #     if dborder.shipping == paramorder.shipping; then return
    #     if dborder.shipping != paramorder.shipping; then (仅仅更新order数据, 无需更新库存信息)
    #
    # update order status
    # if stock(quantity+inflight-preallocation) > 0: 待发货 else 采购
    #
    # 派单需要更新库存占库字段; 重派需要先从之前指派的仓库回滚占库数据, 再更新库存占
    # 库字段;
    #
    # TODO: 需要注意, 订单派单之后, 不能再修改订单jancode, 这个问题后面fix
    #
    def put(self, request, format=None):
        orderInfo = request.data
        logger.debug('派单调试:%s', orderInfo)
        allocate_time = arrow.now().format('YYYY-MM-DD HH:mm:ss')
        paramInventory = orderInfo['inventory']
        relate_inventory = Inventory.objects.get(id=paramInventory)
        # if not paramInventory or not paramShipping:  # 传入参数为空, 无效
        #     return status.HTTP_400_BAD_REQUEST

        # 检查是否订单商品已经在product表中, 如不存在, 返回错误提示
        productObj = None
        try:
            productObj = Product.objects.get(jancode=orderInfo['jancode'])
        except Product.DoesNotExist:
            results = {'errmsg': '商品库中无该商品, 请先创建产品资料'}
            return Response(data=results, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            # 检查被指派的仓库, 是否该产品已经在仓库中存在, 如果不存在, 创建
            stockObj = None
            try:
                stockObj = Stock.objects.get(
                    inventory=paramInventory, product__id=productObj.id)
            except Stock.DoesNotExist:  # 如果第一次分配到该仓库, 主动在该仓库新建产品记录
                stockObj = Stock(
                    product=productObj,
                    inventory=relate_inventory,
                    quantity=0,
                    inflight=0,
                    preallocation=0)
                stockObj.save()
            dborder = Order.objects.get(id=orderInfo['id'])
            dbinventory = dborder.inventory

            rollbackstock = ''
            # 计算库存变化
            if not dbinventory:  # 派单
                stockObj.preallocation = F(
                    'preallocation') + orderInfo['quantity']  # 分配订单需要占库存
            else:  # 重新派单
                if dbinventory.id != paramInventory:  # 重派单, 订单派给了新的仓库, 需要回滚之前的库存占用
                    rollbackstock = Stock.objects.get(
                        inventory=dbinventory.id, product__id=productObj.id)
                    rollbackstock.preallocation = F(
                        'preallocation') - orderInfo['quantity']
                    stockObj.preallocation = F(
                        'preallocation') + orderInfo['quantity']
                else:  # 重派单, 但是仓库没有改变, 无需对库存做更新
                    if dborder.shipping.id == orderInfo[
                            'shipping']:  # 派单信息没有变化, 无需处理.
                        return Response(status=status.HTTP_200_OK)
                    dborder.shipping = Shipping.objects.get(
                        id=orderInfo['shipping'])
                    dborder.save(update_fields=[
                        'shipping',
                    ])
                    return Response(status=status.HTTP_200_OK)

            # 计算订单状态
            stockObj.save()
            # 使用F操作models, save之后需要从数据库刷新, 否则值不能使用
            stockObj.refresh_from_db()

            dborder.allocate_time = allocate_time  # 如果更新库存表, 就需要更新派单时间
            purchaseQuantity = stockObj.preallocation - (
                stockObj.quantity + stockObj.inflight)
            if purchaseQuantity > 0:  # 订单需采购
                if purchaseQuantity < orderInfo['quantity']:
                    dborder.need_purchase = purchaseQuantity
                else:
                    dborder.need_purchase = orderInfo['quantity']
                dborder.status = '待采购'
            else:
                # 到这里, 虽然订单不需要采购, 但是如果在库不能满足发货需求, 该订单需要和
                # 在途的采购单绑定, 同时标记状态为已采购
                if stockObj.preallocation > stockObj.quantity:
                    dborder.status = '已采购'
                    # 找到最新的采购单
                    id = PurchaseOrder.objects.filter(
                        purchaseorderitem__product__jancode=orderInfo[
                            'jancode'],
                        purchaseorderitem__status__isnull=True,
                        status__in=('在途', '部分入库')).aggregate(Max('id'))
                    dborder.purchaseorder = PurchaseOrder.objects.get(
                        id=id['id__max'])
                else:
                    dborder.status = '待发货'

            # 更新订单和仓库信息
            dborder.shipping = Shipping.objects.get(id=orderInfo['shipping'])
            dborder.inventory = relate_inventory
            dborder.save()

            if rollbackstock:
                rollbackstock.save()
            return Response(status=status.HTTP_201_CREATED)
        return Response(status=status.HTTP_400_BAD_REQUEST)


# 在派单页面更新订单需要特殊处理
class OrderAllocateUpdate(views.APIView):
    def post(self, request, format=None):
        updateFields = [
            'jancode',
            'product_title',
            'sku_properties_name',
            'receiver_name',
            'receiver_mobile',
            'receiver_address',
            'quantity',
            'payment',
            'price',
        ]
        orderInfo = request.data
        dbOrderObj = Order.objects.get(id=orderInfo['id'])
        if not dbOrderObj.inventory:  # Not allocation
            for f in updateFields:
                setattr(dbOrderObj, f, orderInfo[f])
            dbOrderObj.save()
        else:  # rollback allocate
            with transaction.atomic():
                productObj = Product.objects.get(jancode=orderInfo['jancode'])
                stockObj = Stock.objects.get(
                    product=productObj, inventory=dbOrderObj.inventory)
                stockObj.preallocation = F(
                    'preallocation') - dbOrderObj.quantity
                stockObj.save()
                for f in updateFields:
                    setattr(dbOrderObj, f, orderInfo[f])
                dbOrderObj.status = '待处理'
                dbOrderObj.need_purchase = None
                dbOrderObj.shipping = None
                dbOrderObj.inventory = None
                dbOrderObj.save()
        return Response(status=status.HTTP_200_OK)


# 把订单拽会到预处理, 回滚派单占用库存(需要考虑购物车订单)
# 1. 根据提交的订单号, 查询所有满足条件的订单(天狗购物车会拆成id-1,id-2),
# 2. 订单状态需要在处理中的状态
# 3. 回滚所有占用的库存
# 4. 不考虑关联采购单,直接设置为空, 保留了订单派单时间
# 5. 设置正确的订单状态
class OrderRollbackToPreprocess(views.APIView):
    def post(self, request, format=None):
        orderid = request.data['orderid']
        dbOrderObjs = Order.objects.filter(
            orderid__contains=orderid,
            status__in=[
                '待处理',
                '待采购',
                '已采购',
                '需介入',
                '待发货',
            ], )
        with transaction.atomic():
            for dbOrderObj in dbOrderObjs:
                productObj = Product.objects.get(jancode=dbOrderObj.jancode)
                try:
                    stockObj = Stock.objects.get(
                        product=productObj, inventory=dbOrderObj.inventory)
                    stockObj.preallocation = F(
                        'preallocation') - dbOrderObj.quantity
                    stockObj.save()
                except Stock.DoesNotExist:
                    continue
                dbOrderObj.status = '待处理'
                dbOrderObj.need_purchase = None
                dbOrderObj.shipping = None
                dbOrderObj.inventory = None
                dbOrderObj.purchaseorder = None
                dbOrderObj.save()

        return Response(status=status.HTTP_200_OK)


# 获取采购列表
class OrderPurchaseList(views.APIView):
    # TODO: 分页
    # 需要返回查询时间, 创建采购单的时候, 我们需要用这个时间来和订单的派单时间做对比, 看看是否能关联相关订单
    # 国内的拼邮单, 采购之前需要看看东京仓库是否有货.
    #
    def get(self, request, format=None):
        sql = "select p.jancode, p.name product_name, p.purchase_link1, p.purchase_link2, p.purchase_link3, p.specification sku_properties_name, min(piad_time) piad_time, max(o.price) product_price, sum(o.need_purchase) qty from stock_product p inner join stock_order o on o.jancode=p.jancode where o.status='待采购' and o.inventory_id=%s group by jancode order by o.id"
        sql2 = "select quantity+inflight-preallocation tokyo_stock from stock_stock where inventory_id='4' and product_id=(select id from stock_product where jancode=%s)"
        sql3 = "select jancode from stock_order where seller_name='天狗' and status='待采购' and jancode=%s"

        def dictfetchall(cursor):
            "Return all rows from a cursor as a dict"
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

        inventory = request.query_params.get('inventory')
        with connection.cursor() as c:
            c.execute(sql, [inventory])
            results = dictfetchall(c)
            for i, r in enumerate(results):
                c.execute(sql3, [r['jancode']])
                isTiangou = c.fetchall()
                results[i]['isTiangou'] = '是' if isTiangou else '否'
                if int(inventory) == 3:  # 需要查东京库存
                    c.execute(sql2, [r['jancode']])
                    rt = c.fetchone()
                    results[i]['tokyo_stock'] = rt[0] if rt else 0
                else:
                    results[i]['tokyo_stock'] = 0
            logger.debug('orderPurchaseList: %s', results)
            data = {
                'data': results,
                'queryTime': arrow.now().format('YYYY-MM-DD HH:mm:ss')
            }
            return Response(data=data, status=status.HTTP_200_OK)
        return Response(status=status.HTTP_400_BAD_REQUEST)


# 生成采购单
class OrderPurchase(views.APIView):
    # TODO: 分页 / 限制只能多采不能少采
    #
    # 流程:
    #   1. 根据注文编号(purchaseorderid)生成purchaseorder.
    #   2. 同时生成purchaseitem.
    #   3. 标记关联订单状态:已采购, 并标注关联purchaseorder. (这里需要考虑采购和派单同时进行情况, 关联订单时需要比较时间)
    #   4. 修改关联库存记录, 增加在途库存.
    #   5. 采购渠道是东京, 占用库存
    #
    def put(self, request, format=None):
        data = request.data.get('data')
        queryTime = request.data.get('queryTime')
        inventory = request.data.get('inventory')
        results = {}
        try:
            with transaction.atomic():
                createtime = arrow.now()
                pos = {}
                for i in data:
                    # create purchaseorder
                    po_id = i.get('purchaseorderid')
                    if not po_id:
                        continue
                    if (not i['quantity'] or i['quantity'] < i['qty'] or
                            not i['supplier'] or not i['price']):
                        results = {'errmsg': '请检查输入'}
                        raise IntegrityError
                    jancode = i['jancode']
                    if po_id not in pos:
                        supplierOb = Supplier.objects.get(id=i['supplier'])
                        inventoryOb = Inventory.objects.get(id=inventory)
                        po = PurchaseOrder(
                            orderid=po_id,
                            supplier=supplierOb,
                            inventory=inventoryOb,
                            create_time=createtime,
                            status='在途', )
                        po.save()
                        pos[po_id] = po

                    # add purchase item
                    price = i['price']
                    poitem = PurchaseOrderItem(
                        product=Product.objects.get(jancode=jancode),
                        quantity=i['quantity'],
                        purchaseorder=pos[po_id],
                        price=price)
                    poitem.save()

                    # set inflight in stock
                    stock = Stock.objects.get(
                        inventory=inventory,
                        product__jancode=jancode, )
                    stock.inflight = F('inflight') + int(i['quantity'])
                    stock.save()

                    # preallocation stock if supplier is tokyo
                    if '东京仓' in supplierOb.name:
                        stockTokyo = Stock.objects.get(
                            inventory=Inventory.objects.get(name='东京'),
                            product__jancode=jancode, )
                        stockTokyo.preallocation = F('preallocation') + int(
                            i['quantity'])
                        stockTokyo.save()

                    # 1. update order, notes: may qty != quantity
                    # 2. if allocate_time > querytime, don't relation it.
                    orders = Order.objects.filter(
                        inventory=inventory, jancode=jancode, status='待采购')
                    # orders_qty_sum = 0
                    for o in orders:
                        # django存的是naive的时间, 所以我们这里也要用才能比较
                        if o.allocate_time > arrow.get(queryTime).naive:
                            continue
                        o.purchaseorder = pos[po_id]
                        o.status = '已采购'
                        o.save(update_fields=['status', 'purchaseorder'])
                        # orders_qty_sum += o.need_purchase
                        # if orders_qty_sum > int(i['quantity']):
                        #     break
        except IntegrityError:
            return Response(data=results, status=status.HTTP_400_BAD_REQUEST)

        return Response(status=status.HTTP_201_CREATED)


# 提交采购单(直接采购, 不参考待采购列表)
class NoOrderPurchase(views.APIView):
    #
    # 流程:
    #   1. 根据注文编号(orderid)生成purchaseorder.
    #   2. 同时生成purchaseitem.
    #   3. 标记关联订单状态:已采购, 并标注关联purchaseorder ( 如果有的话, 需要关联, 可以减少重复采购 ).
    #   4. 修改关联库存记录, 增加在途库存.
    #   postForm: {
    #   inventory: undefined,
    #   supplier: undefined,
    #   orderid: '',
    #   items: [
    #     {
    #       jancode: undefined,
    #       quantity: undefined,
    #       price: undefined
    #     }
    #   ]
    # },
    #
    def put(self, request, format=None):
        data = request.data
        inventory = data['inventory']
        results = {}
        try:
            with transaction.atomic():
                createtime = arrow.now()
                # create purchaseorder
                supplierObj = Supplier.objects.get(id=data['supplier'])
                inventoryObj = Inventory.objects.get(id=inventory)
                poObj = PurchaseOrder(
                    orderid=data['orderid'],
                    supplier=supplierObj,
                    inventory=inventoryObj,
                    create_time=createtime,
                    status='在途', )
                poObj.save()
                for i in data['items']:
                    # add purchase item
                    try:
                        productObj = Product.objects.get(jancode=i['jancode'])
                    except Product.DoesNotExist:
                        results = {
                            'errmsg': '商品库中无该商品%s, 请先创建产品资料' % (i['jancode'], )
                        }
                        raise IntegrityError
                    poitemObj = PurchaseOrderItem(
                        product=productObj,
                        quantity=i['quantity'],
                        purchaseorder=poObj,
                        price=i['price'])
                    poitemObj.save()

                    try:
                        stockObj = Stock.objects.get(
                            inventory=inventoryObj,
                            product__jancode=i['jancode'])
                    except Stock.DoesNotExist:  # 如果第一次分配到该仓库, 主动在该仓库新建产品记录
                        stockObj = Stock(
                            product=productObj,
                            inventory=inventoryObj,
                            quantity=0,
                            inflight=0,
                            preallocation=0)
                    stockObj.inflight = stockObj.inflight + int(i['quantity'])
                    stockObj.save()

                    # preallocation stock if supplier is tokyo
                    if '东京仓' in supplierObj.name:
                        stockTokyo = Stock.objects.get(
                            inventory=Inventory.objects.get(name='东京'),
                            product__jancode=i['jancode'], )
                        stockTokyo.preallocation = F('preallocation') + int(
                            i['quantity'])
                        stockTokyo.save()

                    orders = Order.objects.filter(
                        inventory__id=inventory,
                        jancode=i['jancode'],
                        status='待采购').order_by('id')
                    c = int(i['quantity'])
                    for o in orders:
                        if o.need_purchase > c:
                            break
                        o.purchaseorder = poObj
                        o.status = '已采购'
                        o.save(update_fields=['status', 'purchaseorder'])
                        c = c - o.need_purchase
                return Response(status=status.HTTP_201_CREATED)
        except IntegrityError as e:
            if e.args and e.args[0] == 1062:
                return Response(
                    data={'errmsg': '同一产品需合并录入'},
                    status=status.HTTP_400_BAD_REQUEST)
            return Response(data=results, status=status.HTTP_400_BAD_REQUEST)


# 删除采购单
class PurchaseOrderDelete(views.APIView):
    # TODO: 增加是否需要重新采购的选项给用户
    #
    # 流程:
    #   1. 标记采购单状态: 删除.
    #   2. 取消关联订单的采购信息, 并重新标识订单状态为待采购.
    #   3. 修改关联库存记录, 减少在途库存.
    #
    #   request param: id
    def put(self, request, format=None):
        id = request.data.get('id')
        poObj = PurchaseOrder.objects.get(id=id)  # 采购单
        poitemObjs = poObj.purchaseorderitem.all()  # 关联采购商品
        orderObjs = poObj.order.filter(status='已采购')  # 关联订单, 状态为已采购

        with transaction.atomic():

            # rollback order, set order status
            for o in orderObjs:
                o.status = '待采购'
                o.purchaseorder = None  # clear relate po
                o.save(update_fields=['status', 'purchaseorder'])

            # rollback stock, set stock preallocation
            for poi in poitemObjs:
                stockObj = Stock.objects.get(
                    product=poi.product, inventory=poObj.inventory)
                stockObj.inflight = F('inflight') - poi.quantity
                stockObj.save(update_fields=[
                    'inflight',
                ])

                # rollback tokyo stock if supplier is tokyo
                if '东京仓' in poObj.supplier.name:
                    stockTokyo = Stock.objects.get(
                        inventory=Inventory.objects.get(name='东京'),
                        product=poi.product, )
                    stockTokyo.preallocation = F('preallocation') - poi.quantity
                    stockTokyo.save()

            # mark purchaseorder status as '删除'
            poObj.status = '删除'
            poObj.save(update_fields=['status'])
            return Response(status=status.HTTP_200_OK)

        return Response(status=status.HTTP_400_BAD_REQUEST)


# 清采购单(采购单入库)
class PurchaseOrderClear(views.APIView):
    # 流程:
    #   1. 标记采购单状态: 已入库.
    #   2. stock入库(减inflight, 增加quantity).
    #   3. 标记订单待发货.
    # 需要支持部分入库
    # 采购流程中, 国内采购需要增加转运, 流程如下:
    #   1. 采购商品到达东京, 东京仓库入库, 这个入库只是标记purchaseorderitem状态为"东京仓", 不做其他操作
    #   2. 增加页面, 显示所有状态为东京仓的purchaseorderitem, 东京仓库根据这个页面转运商品, 关联转运单号
    #   3. 广州仓人员收到转运包括, 做入库操作
    def put(self, request, format=None):
        id = request.data.get('id')
        inventory_id = request.data.get('inventory')
        pois = request.data.get('pois')
        poObj = PurchaseOrder.objects.get(id=id)  # 采购单

        with transaction.atomic():
            # mark purchaseorder
            # poObj.status = '入库'
            # poObj.save(update_fields=['status'])

            # stock in
            for poi in pois:
                poiObj = poObj.purchaseorderitem.get(
                    product__jancode=poi['jancode'])
                if not poi['qty'] or '已入库' == poiObj.status:
                    continue

                if inventory_id == 3:  # 如果是广州仓库, 只标记订单明细状态
                    if poiObj.status in ['东京仓', '转运中']:
                        continue
                    poiObj.status = '东京仓'
                    poiObj.save()
                    continue

                poiObj.status = '已入库'
                poiObj.save()

                inventory = Inventory.objects.get(id=inventory_id)
                stockObj = Stock.objects.get(
                    product__jancode=poi['jancode'], inventory=inventory)
                # poi.quantity记录的是采购数量, qty是实际到库数量.
                # 入库实际到库数量, 扣减inflight数量用采购数量.
                # TODO: 如果实际到库少于采购数量, 需要处理漏采(漏采需补采购)
                if poiObj.quantity < poi['qty']:
                    stockObj.quantity = F('quantity') + poi['qty']
                else:
                    stockObj.quantity = F('quantity') + poiObj.quantity
                stockObj.inflight = F('inflight') - poiObj.quantity
                stockObj.save()

                # tokyo stock out if supplier is tokyo
                if '东京仓' in poObj.supplier.name:
                    stockTokyo = Stock.objects.get(
                        inventory=Inventory.objects.get(name='东京'),
                        product__jancode=poi['jancode'], )
                    stockTokyo.preallocation = F(
                        'preallocation') - poiObj.quantity
                    if poiObj.quantity < poi['qty']:
                        stockTokyo.quantity = F('quantity') - poi['qty']
                    else:
                        stockTokyo.quantity = F('quantity') - poiObj.quantity
                    stockTokyo.save()

                poObj.order.filter(
                    jancode=poi['jancode'], status='已采购').update(status='待发货')

            count = poObj.purchaseorderitem.filter(status='已入库').count()
            if count > 0:
                if count != len(pois):
                    poObj.status = '部分入库'
                else:
                    poObj.status = '入库'
                poObj.save(update_fields=[
                    'status',
                ])
            return Response(status=status.HTTP_200_OK)

        return Response(status=status.HTTP_400_BAD_REQUEST)


class PurchaseOrderTransform(views.APIView):
    # 标记purchaseorderitem状态, 记录转运单号
    def post(self, request, format=None):
        data = request.data
        logger.debug('purchaseOrderTransform debug: %s', data)
        with transaction.atomic():
            for i in data['purchaseorderitems']:
                poiObj = PurchaseOrderItem.objects.get(id=i['id'])
                poiObj.status = '转运中'
                poiObj.delivery_no = data['delivery_no']
                poiObj.save()
        return Response(status=status.HTTP_200_OK)


class DomesticStockIn(views.APIView):
    def post(self, request, format=None):
        data = request.data
        with transaction.atomic():
            poiObj = PurchaseOrderItem.objects.get(id=data['id'])
            poObj = poiObj.purchaseorder
            if '转运中' in poiObj.status:
                poiObj.status = '已入库'
                poiObj.save()

                inventory = poObj.inventory
                stockObj = Stock.objects.get(
                    product=poiObj.product, inventory=inventory)
                stockObj.quantity = F('quantity') + poiObj.quantity
                stockObj.inflight = F('inflight') - poiObj.quantity
                stockObj.save()

                # tokyo stock out if supplier is tokyo
                if '东京仓' in poObj.supplier.name:
                    stockTokyo = Stock.objects.get(
                        inventory=Inventory.objects.get(name='东京'),
                        product=poiObj.product, )
                    stockTokyo.preallocation = F(
                        'preallocation') - poiObj.quantity
                    stockTokyo.quantity = F('quantity') - poiObj.quantity
                    stockTokyo.save()

                poObj.order.filter(
                    jancode=poiObj.product.jancode, status='已采购').update(
                        status='待发货')

            count = poObj.purchaseorderitem.filter(status='已入库').count()
            total = poObj.purchaseorderitem.all().count()
            if count > 0:
                if count != total:
                    poObj.status = '部分入库'
                else:
                    poObj.status = '入库'
                poObj.save(update_fields=[
                    'status',
                ])
            return Response(status=status.HTTP_200_OK)


# 采购标记订单疑难
class OrderMarkConflict(views.APIView):
    # TODO: 分页
    #
    # 流程:
    #   1. 修改订单状态为: 需介入.
    #   2. 标记订单状态conflict字段
    #
    def put(self, request, format=None):
        data = request.data

        with transaction.atomic():
            for i in data:
                # mark conflict
                shipping = Shipping.objects.get(id=i['shipping'])
                inventory = Inventory.objects.get(id=i['inventory'])
                i.pop('shipping_name')
                i.pop('inventory_name')
                i.pop('db_number')
                i.pop('purchaseorder_orderid')
                i.pop('shippingdb_delivery_time')
                i['status'] = '需介入'
                i['shipping'] = shipping
                i['inventory'] = inventory
                i['purchaseorder'] = None
                s = Order(**i)
                s.save()
            return Response(status=status.HTTP_200_OK)

        return Response(status=status.HTTP_400_BAD_REQUEST)


# 需介入处理(协调退换货)
# 退款: 设置订单状态:已删除, 释放占用的库存. 换货: 释放占用的库存, 重新占用新库存.
# 注意: 换货意味着重新派单, 需要设置派单时间
class OrderConflict(views.APIView):
    def put(self, request, format=None):
        data = request.data

        with transaction.atomic():
            if '已删除' in data['status']:  # 退款
                stockObj = Stock.objects.get(
                    inventory=data['inventory'],
                    product__jancode=data['jancode'])
                stockObj.preallocation = F('preallocation') - data['quantity']
                stockObj.save()
                orderObj = Order.objects.get(id=data['id'])
                orderObj.status = '已删除'
                orderObj.conflict_feedback = data['conflict_feedback']
                orderObj.save(update_fields=['status', 'conflict_feedback'])
            else:  # 更换
                orderObj = Order.objects.get(id=data['id'])
                if orderObj.jancode != data['jancode']:
                    # 判断新jancode库存是否满足
                    stockObj = None
                    try:
                        stockObj = Stock.objects.get(
                            inventory=data['inventory'],
                            product__jancode=data['jancode'])
                    except Stock.DoesNotExist:  # 如果第一次分配到该仓库, 主动在该仓库新建产品记录
                        if not Product.objects.filter(
                                jancode=data['jancode']).exists():
                            errmsg = {'errmsg': '商品库中无该商品, 请先创建产品资料'}
                            return Response(
                                data=errmsg,
                                status=status.HTTP_400_BAD_REQUEST)
                        stockObj = Stock(
                            product=Product.objects.get(
                                jancode=data['jancode']),
                            inventory=Inventory.objects.get(
                                id=data['inventory']),
                            quantity=0,
                            inflight=0,
                            preallocation=0)
                        stockObj.save()

                    # 回滚旧的jancode库存占用
                    rollbackStockObj = Stock.objects.get(
                        inventory=orderObj.inventory,
                        product__jancode=orderObj.jancode)
                    rollbackStockObj.preallocation = F(
                        'preallocation') - data['quantity']
                    rollbackStockObj.save()

                    real_stock_qty = stockObj.quantity + stockObj.inflight - stockObj.preallocation
                    if data['quantity'] <= real_stock_qty:  # 库存足够, 记得取消采购标记(need_purchase=0)
                        stockObj.preallocation = F(
                            'preallocation') + data['quantity']
                        orderObj.status = '待发货'
                        orderObj.need_purchase = 0
                        orderObj.jancode = data['jancode']
                    else:
                        need_purchase = (
                            data['quantity'] - real_stock_qty
                        ) if real_stock_qty > 0 else data['quantity']
                        orderObj.need_purchase = need_purchase
                        orderObj.jancode = data['jancode']
                        orderObj.status = '待采购'
                        stockObj.preallocation = F(
                            'preallocation') + data['quantity']
                    orderObj.allocate_time = arrow.now().format(
                        'YYYY-MM-DD HH:mm:ss')  # 需要更新订单分配时间
                    orderObj.conflict_feedback = data['conflict_feedback']
                    orderObj.save()
                    stockObj.save()
                else:  # 用户没有做任何操作, 直接改订单状态为待采购
                    orderObj.status = '待采购'
                    orderObj.conflict_feedback = data['conflict_feedback']
                    orderObj.save(
                        update_fields=['status', 'conflict_feedback'])

            return Response(status=status.HTTP_200_OK)

        return Response(status=status.HTTP_400_BAD_REQUEST)


# 订单删除流程:
#   if status == '预分配' then mark status=已删除
#   elif status == '待发货'|'需介入'|'已采购'|'待采购' then
#       mark order status = 已删除
#       rollback stock
#   else then
#       return error
#
class OrderDelete(views.APIView):
    def put(self, request, format=None):
        data = request.data

        orderObj = Order.objects.get(id=data['id'])
        with transaction.atomic():
            if '待处理' in orderObj.status:  # 直接标记删除
                orderObj.status = '已删除'
                orderObj.conflict_feedback = data['conflict_feedback']
                orderObj.save()
            elif ('待发货' in orderObj.status or '需介入' in orderObj.status or
                  '已采购' in orderObj.status or
                  '待采购' in orderObj.status):  # 清除占用的库存
                stockObj = Stock.objects.get(
                    inventory=orderObj.inventory,
                    product__jancode=orderObj.jancode)
                stockObj.preallocation = F('preallocation') - orderObj.quantity
                stockObj.save()
                orderObj.status = '已删除'
                orderObj.conflict_feedback = data['conflict_feedback']
                orderObj.save()
            elif '已删除' in orderObj.status:
                pass
            else:  # 不可删除, 已经分配DB单号, 需要先删除DB单号, 或者已经发货, 无法删除
                errmsg = {'errmsg': '订单已分配面单, 无法删除'}
                return Response(
                    data=errmsg, status=status.HTTP_400_BAD_REQUEST)

            return Response(status=status.HTTP_200_OK)

        return Response(status=status.HTTP_400_BAD_REQUEST)


# 更新产品包括条码
# 更新条码需要同步更新库存表和订单表
class ProductUpdateJancode(views.APIView):
    def post(self, request, format=None):
        data = request.data

        with transaction.atomic():
            productObj = Product.objects.get(id=data['id'])
            jancode = productObj.jancode
            productSerializer = ProductSerializer(
                productObj, data=request.data)
            if productSerializer.is_valid():
                productSerializer.save()
                Order.objects.filter(jancode=jancode).update(
                    jancode=data['jancode'],
                    product_title=productObj.name,
                    sku_properties_name=productObj.specification, )
                return Response(status=status.HTTP_200_OK)

        return Response(status=status.HTTP_400_BAD_REQUEST)


# 采购单入库: 需要考虑对应订单被删除的情况


class UserInfo(generics.ListAPIView):
    queryset = Token.objects.all()
    serializer_class = TokenSerializer
    filter_backends = (django_filters.rest_framework.DjangoFilterBackend, )
    filter_fields = ('key', )


class PurchaseOrderList(generics.ListCreateAPIView):

    # def get_queryset(self):
    #     queryset = PurchaseOrder.objects.all()
    #     workspace = self.request.query_params.get('workspace')
    #     airline = self.request.query_params.get('airline')

    #     if workspace:
    #         queryset = queryset.filter(workspace_id=workspace)
    #     elif airline:
    #         queryset = queryset.filter(workspace__airline_id=airline)

    #     return queryset

    queryset = PurchaseOrder.objects.all()
    serializer_class = PurchaseOrderSerializer
    filter_backends = (django_filters.rest_framework.DjangoFilterBackend, )
    # filter_fields = ('orderid', 'inventory', 'supplier', 'status')
    filter_class = PurchaseOrderFilter


class PurchaseOrderDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = PurchaseOrder.objects.all()
    serializer_class = PurchaseOrderSerializer


class PurchaseOrderItemList(generics.ListCreateAPIView):
    queryset = PurchaseOrderItem.objects.all()
    serializer_class = PurchaseOrderItemSerializer
    filter_backends = (django_filters.rest_framework.DjangoFilterBackend, )
    filter_class = PurchaseOrderItemFilter


class PurchaseOrderItemDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = PurchaseOrderItem.objects.all()
    serializer_class = PurchaseOrderItemSerializer


class ProductList(generics.ListCreateAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    filter_backends = (django_filters.rest_framework.DjangoFilterBackend, )
    # filter_fields = ('jancode', 'category', 'name', 'brand')
    filter_class = ProductFilter


class ProductDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer


class BondedProductList(generics.ListCreateAPIView):
    queryset = BondedProduct.objects.all()
    serializer_class = BondedProductSerializer
    filter_backends = (django_filters.rest_framework.DjangoFilterBackend, )
    filter_fields = ('jancode', 'bonded_name')


class BondedProductDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = BondedProduct.objects.all()
    serializer_class = BondedProductSerializer


class OrderList(generics.ListCreateAPIView):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer
    filter_backends = (django_filters.rest_framework.DjangoFilterBackend, )
    filter_class = OrderFilter
    # filter_fields = ('orderid', 'channel_name', 'receiver_name', 'jancode',
    #                  'status')


class OrderDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer


class StockList(generics.ListCreateAPIView):
    queryset = Stock.objects.all()
    serializer_class = StockSerializer
    filter_backends = (django_filters.rest_framework.DjangoFilterBackend, )
    filter_class = StockFilter


class StockDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = Stock.objects.all()
    serializer_class = StockSerializer


class ShippingList(generics.ListCreateAPIView):
    queryset = Shipping.objects.all()
    serializer_class = ShippingSerializer
    filter_backends = (django_filters.rest_framework.DjangoFilterBackend, )
    filter_fields = ('inventory', )


class ShippingDBList(generics.ListAPIView):
    queryset = ShippingDB.objects.all()
    serializer_class = ShippingDBSerializer
    filter_backends = (django_filters.rest_framework.DjangoFilterBackend, )
    filter_class = ShippingDBFilter


class InventoryList(generics.ListCreateAPIView):
    queryset = Inventory.objects.all()
    serializer_class = InventorySerializer
    filter_backends = (django_filters.rest_framework.DjangoFilterBackend, )


class SupplierList(generics.ListCreateAPIView):
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    filter_backends = (django_filters.rest_framework.DjangoFilterBackend, )
